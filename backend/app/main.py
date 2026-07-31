from fastapi import FastAPI, Depends, HTTPException, status, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text
from datetime import date, datetime
from pydantic import BaseModel, EmailStr
from typing import Optional
from pathlib import Path
import httpx

from app.config import settings
from app.database import init_db, get_db
from app.models import User, Licitacion, SubscriptionPlan, ExtractionLog, KeywordAlert, PipelineItem, ScheduledReport
from app.auth import hash_password, verify_password, create_token, get_current_user

RECURRENTE_API = "https://app.recurrente.com/api"
RECURRENTE_PLANS = {
    "basico":  {"price_id": "price_lltzdrus", "name": "B\u00e1sico",  "price": 349, "keywords": 10,  "pipeline": 3,  "schedules": 1, "users": 1},
    "pro":     {"price_id": "price_kyqlcwp6", "name": "Pro",        "price": 599, "keywords": 50,  "pipeline": 30, "schedules": 5, "users": 3},
    "enterprise": {"price_id": "price_n2pdn7xh", "name": "Enterprise","price": 999, "keywords": 999, "pipeline": 999, "schedules": 999, "users": 10},
}
PIPELINE_ETAPAS = ["deteccion", "analisis", "preparacion", "presentacion", "adjudicacion", "ganada", "perdida"]

def _owner_id(user: User) -> int:
    return user.main_user_id or user.id

def _send_invite_email(to_email: str, name: str, owner_email: str, temp_password: str | None):
    import asyncio
    url = settings.FRONTEND_URL.strip('/') if settings.FRONTEND_URL else 'https://licitrackgt-production.up.railway.app'
    async def _do():
        try:
            from app.services.email_service import enviar_correo
            html = f"""<div style="font-family:Arial;max-width:600px;margin:auto;color:#222">
                <h2 style="color:#1a3a5c">Bienvenido a LiciTrackGT</h2>
                <p>Hola {name or 'usuario'},</p>
                <p>{owner_email} te ha invitado a colaborar en <b>LiciTrackGT</b>.</p>
                {f'<p>Tu contraseña temporal: <b style="font-size:18px;background:#f0f0f0;padding:4px 8px;border-radius:4px">{temp_password}</b></p>' if temp_password else '<p>Ya tienes una cuenta. Tu acceso ha sido actualizado.</p>'}
                <p>Ingresa en: <a href="{url}" style="color:#1a5fb4">{url}</a></p>
                <p style="margin-top:16px">Desde aqui podras buscar licitaciones, configurar alertas y dar seguimiento a tus oportunidades.</p>
                <p style="color:#888;font-size:12px">LiciTrackGT - Monitoreo inteligente de Guatecompras</p></div>"""
            await enviar_correo([to_email], "Has sido invitado a LiciTrackGT", html)
        except Exception as e:
            print(f"Error enviando invitacion a {to_email}: {e}")
    asyncio.create_task(_do())

app = FastAPI(title="LiciTrackGT API", docs_url="/docs")

app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ============================================================
# SCHEMAS
# ============================================================
class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict

class RegisterRequest(BaseModel):
    email: EmailStr
    password: str
    name: str = ""
    whatsapp_phone: Optional[str] = None

# ============================================================
# AUTH ROUTES
# ============================================================
@app.post("/api/auth/login")
async def login(req: LoginRequest, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.email == req.email))
    user = result.scalar_one_or_none()
    if not user or not verify_password(req.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Credenciales invalidas")
    token = create_token({"sub": user.email})
    return TokenResponse(access_token=token, user={
        "id": user.id, "email": user.email, "name": user.name,
        "is_admin": user.is_admin, "plan": user.subscription_plan,
        "is_team_member": user.main_user_id is not None
    })

@app.post("/api/auth/register")
async def register(req: RegisterRequest, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.email == req.email))
    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Email ya registrado")
    user = User(email=req.email, password_hash=hash_password(req.password), name=req.name)
    if req.whatsapp_phone:
        user.whatsapp_phone = req.whatsapp_phone.strip()
    db.add(user); await db.commit(); await db.refresh(user)
    token = create_token({"sub": user.email})
    return TokenResponse(access_token=token, user={
        "id": user.id, "email": user.email, "name": user.name, "plan": "free"
    })

class FirebaseAuthRequest(BaseModel):
    firebase_token: str
    name: str = ""

@app.post("/api/auth/firebase")
async def firebase_auth(req: FirebaseAuthRequest, db: AsyncSession = Depends(get_db)):
    try:
        import httpx
        async with httpx.AsyncClient(timeout=15) as cl:
            resp = await cl.get(
                "https://identitytoolkit.googleapis.com/v1/accounts:lookup",
                params={"key": "AIzaSyBtdzASSqHz2oirxJGl6deGkfIUBMUnO_c"},
                json={"idToken": req.firebase_token},
            )
        if resp.status_code != 200:
            raise HTTPException(status_code=401, detail="Token Firebase invalido")
        data = resp.json()
        users = data.get("users", [])
        if not users:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        fb_user = users[0]
        email = fb_user.get("email", "")
        if not email:
            email = fb_user.get("providerUserInfo", [{}])[0].get("email", "")
        name = req.name or fb_user.get("displayName", "") or email.split("@")[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Error verificando Firebase: {str(e)[:100]}")
    
    result = await db.execute(select(User).where(User.email == email))
    user = result.scalar_one_or_none()
    if not user:
        user = User(email=email, password_hash=hash_password("firebase-no-password"), name=name)
        db.add(user); await db.commit(); await db.refresh(user)
    token = create_token({"sub": user.email})
    return TokenResponse(access_token=token, user={
        "id": user.id, "email": user.email, "name": user.name,
        "is_admin": user.is_admin, "plan": user.subscription_plan,
        "is_team_member": user.main_user_id is not None
    })

@app.get("/api/auth/me")
async def me(user: User = Depends(get_current_user)):
    return {"id": user.id, "email": user.email, "name": user.name,
            "is_admin": user.is_admin, "plan": user.subscription_plan,
            "status": user.subscription_status, "whatsapp_phone": user.whatsapp_phone or "",
            "is_team_member": user.main_user_id is not None, "owner_id": user.main_user_id}

class UpdateProfileRequest(BaseModel):
    name: Optional[str] = None
    whatsapp_phone: Optional[str] = None
    current_password: Optional[str] = None
    new_password: Optional[str] = None

@app.put("/api/auth/profile")
async def update_profile(req: UpdateProfileRequest, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if req.name is not None: user.name = req.name
    if req.whatsapp_phone is not None: user.whatsapp_phone = req.whatsapp_phone.strip() or None
    if req.new_password:
        if not req.current_password or not verify_password(req.current_password, user.password_hash):
            raise HTTPException(status_code=400, detail="Contraseña actual incorrecta")
        user.password_hash = hash_password(req.new_password)
    await db.commit()
    return {"ok": True, "whatsapp_phone": user.whatsapp_phone or "", "password_changed": bool(req.new_password)}

class TestNotificationRequest(BaseModel):
    whatsapp_phone: Optional[str] = None

@app.post("/api/auth/test-notification")
async def test_notification(req: TestNotificationRequest, user: User = Depends(get_current_user)):
    from app.services.email_service import enviar_correo
    from app.services.whatsapp_service import enviar_whatsapp
    result = {"email": False, "whatsapp": False}
    try:
        await enviar_correo([user.email], "LiciTrackGT - Prueba de notificaciones",
            '<div style="font-family:Arial;max-width:600px;margin:auto"><h2 style="color:#1a3a5c">LiciTrackGT</h2>'
            '<p>Esta es una <b>prueba</b> de que las notificaciones por correo funcionan correctamente.</p>'
            '<p style="color:green;font-weight:bold">Correo configurado exitosamente.</p>'
            '<p style="color:#888;font-size:12px">Recibiras tus alertas y reportes en esta direccion.</p></div>')
        result["email"] = True
    except Exception as e:
        result["email_error"] = str(e)[:100]
    phone = req.whatsapp_phone or user.whatsapp_phone
    if phone:
        try:
            ok = await enviar_whatsapp(phone,
                "LiciTrackGT - Prueba de WhatsApp\n\n"
                "Este es un mensaje de prueba para confirmar que recibiras tus alertas por aqui.\n\n"
                "Cuando haya nuevas licitaciones que coincidan con tus keywords, te llegara un aviso como este.")
            result["whatsapp"] = ok
            if not ok:
                result["whatsapp_error"] = "No se pudo enviar (verifica token/número)"
        except Exception as e:
            result["whatsapp_error"] = str(e)[:100]
    else:
        result["whatsapp_error"] = "No has configurado un numero de WhatsApp"
    return result

@app.api_route("/api/whatsapp/webhook", methods=["GET", "POST"], include_in_schema=False)
async def whatsapp_webhook(request: Request):
    from app.services.whatsapp_service import verificar_webhook
    if request.method == "GET":
        params = request.query_params
        challenge = verificar_webhook(
            params.get("hub.mode", ""), params.get("hub.verify_token", ""),
            params.get("hub.challenge", "")
        )
        if challenge:
            from fastapi.responses import PlainTextResponse
            return PlainTextResponse(challenge, status_code=200)
        raise HTTPException(status_code=403, detail="Token invalido")
    else:
        import json
        body = await request.body()
        try:
            data = json.loads(body)
            print(f"WhatsApp webhook recibido: {json.dumps(data, indent=2)[:300]}")
        except:
            pass
        return {"ok": True}

# ============================================================
# DASHBOARD
# ============================================================
@app.get("/api/dashboard/stats")
async def dashboard_stats(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    total = await db.scalar(select(func.count(Licitacion.id)))
    entidades = await db.scalar(select(func.count(func.distinct(Licitacion.entidad_compradora))))
    monto_prom = await db.scalar(select(func.avg(Licitacion.monto)).where(Licitacion.monto > 0)) or 0
    total_2026 = await db.scalar(select(func.count(Licitacion.id)).where(Licitacion.anio == 2026)) or 0
    top_entidades = (await db.execute(
        select(Licitacion.entidad_compradora, func.count().label("cnt"))
        .where(Licitacion.entidad_compradora != "")
        .group_by(Licitacion.entidad_compradora).order_by(text("cnt DESC")).limit(15)
    )).all()
    top_categorias = (await db.execute(
        select(Licitacion.categoria, func.count().label("cnt"))
        .where(Licitacion.categoria != "")
        .group_by(Licitacion.categoria).order_by(text("cnt DESC")).limit(10)
    )).all()
    meses_2026 = (await db.execute(
        select(Licitacion.mes, func.count().label("cnt"))
        .where(Licitacion.anio == 2026).group_by(Licitacion.mes).order_by(Licitacion.mes)
    )).all()
    por_mes = [{"mes": m, "cantidad": c} for m, c in meses_2026]
    por_departamento = (await db.execute(
        select(Licitacion.departamento, func.count().label("cnt"))
        .where(Licitacion.departamento != "").group_by(Licitacion.departamento)
        .order_by(text("cnt DESC")).limit(10)
    )).all()
    return {
        "total": total or 0, "entidades": entidades or 0,
        "monto_prom": round(float(monto_prom), 2), "total_2026": total_2026 or 0,
        "entidades_top": [{"nombre": r[0], "cantidad": r[1]} for r in top_entidades],
        "categorias_top": [{"nombre": r[0], "cantidad": r[1]} for r in top_categorias],
        "por_mes": por_mes,
        "por_departamento": [{"nombre": r[0], "cantidad": r[1]} for r in por_departamento],
    }

# ============================================================
# LICITACIONES
# ============================================================
class FiltrosQuery(BaseModel):
    estatus: Optional[str] = None
    categoria: Optional[str] = None
    entidad: Optional[str] = None
    texto: Optional[str] = None
    anio: Optional[int] = None
    mes: Optional[int] = None
    departamento: Optional[str] = None
    metodo: Optional[str] = None
    modalidad: Optional[str] = None
    fecha_desde: Optional[str] = None
    fecha_hasta: Optional[str] = None
    monto_min: Optional[float] = None
    monto_max: Optional[float] = None
    destinatario: Optional[str] = None
    destinatarios: Optional[str] = None
    page: int = 1
    per_page: int = 50

def _apply_filtros(q, f):
    if f.estatus: q = q.where(Licitacion.estado == f.estatus)
    if f.categoria: q = q.where(Licitacion.categoria == f.categoria)
    if f.entidad: q = q.where(Licitacion.entidad_compradora.ilike(f"%{f.entidad}%"))
    if f.texto: q = q.where(Licitacion.titulo.ilike(f"%{f.texto}%"))
    if f.anio: q = q.where(Licitacion.anio == f.anio)
    if f.mes: q = q.where(Licitacion.mes == f.mes)
    if f.departamento: q = q.where(Licitacion.departamento == f.departamento)
    if f.metodo: q = q.where(Licitacion.metodo == f.metodo)
    if f.modalidad: q = q.where(Licitacion.modalidad.ilike(f"%{f.modalidad}%"))
    if f.fecha_desde: q = q.where(Licitacion.fecha_publicacion >= f.fecha_desde)
    if f.fecha_hasta: q = q.where(Licitacion.fecha_publicacion <= f.fecha_hasta)
    if f.monto_min: q = q.where(Licitacion.monto >= f.monto_min)
    if f.monto_max: q = q.where(Licitacion.monto <= f.monto_max)
    return q

def _serialize(r):
    return [r.nog, r.ocid, str(r.fecha_publicacion) if r.fecha_publicacion else "",
            r.titulo or "", r.entidad_compradora or "", r.monto or 0, r.moneda or "GTQ",
            r.estado or "", r.categoria or "", r.metodo or "", r.modalidad or "",
            r.departamento or ""]

CSV_HEADERS = ["NOG", "OCID", "Fecha", "Titulo", "Entidad", "Monto", "Moneda",
               "Estado", "Categoria", "Metodo", "Modalidad", "Departamento"]

def _relevance(titulo: str, keyword: str) -> int:
    if not keyword: return 0
    tl = titulo.lower()
    kw = keyword.lower()
    score = 0
    parts = [k.strip() for k in kw.replace(",", " ").replace(";", " ").split() if k.strip()]
    if not parts: parts = [kw]
    for p in parts:
        if p in tl: score += 100 // max(1, len(parts))
    return min(100, score)

@app.post("/api/licitaciones")
async def query_licitaciones(f: FiltrosQuery, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    q = _apply_filtros(select(Licitacion), f)
    total = await db.scalar(select(func.count()).select_from(q.subquery()))
    offset = (f.page - 1) * f.per_page
    q = q.order_by(Licitacion.fecha_publicacion.desc()).offset(offset).limit(f.per_page)
    rows = (await db.execute(q)).scalars().all()
    keyword_text = f.texto.lower().strip() if f.texto else ""
    return {"total": total or 0, "page": f.page, "data": [{
        "nog": r.nog, "ocid": r.ocid, "fecha": str(r.fecha_publicacion) if r.fecha_publicacion else "",
        "titulo": r.titulo, "entidad": r.entidad_compradora, "monto": r.monto or 0,
        "estado": r.estado, "categoria": r.categoria, "metodo": r.metodo,
        "modalidad": r.modalidad, "departamento": r.departamento,
        "relevancia": _relevance(r.titulo or "", keyword_text) if keyword_text else None,
    } for r in rows]}

@app.post("/api/licitaciones/export")
async def export_licitaciones(f: FiltrosQuery, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    import io, csv
    from fastapi.responses import StreamingResponse
    q = _apply_filtros(select(Licitacion), f)
    base = q.order_by(Licitacion.id)

    async def rows_batches(batch=2000):
        last_id = 0
        while True:
            sub = base.where(Licitacion.id > last_id).limit(batch)
            batch_rows = (await db.execute(sub)).scalars().all()
            if not batch_rows:
                break
            for r in batch_rows:
                yield r
            last_id = batch_rows[-1].id

    async def stream_csv():
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(CSV_HEADERS)
        yield buf.getvalue()
        buf = io.StringIO()
        writer = csv.writer(buf)
        async for r in rows_batches():
            writer.writerow(_serialize(r))
            if buf.tell() > 1024 * 512:
                yield buf.getvalue()
                buf = io.StringIO()
                writer = csv.writer(buf)
        yield buf.getvalue()

    filename = f"licitaciones_{f.anio or 'todos'}_{f.mes or 'todos'}.csv"
    return StreamingResponse(stream_csv(), media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.post("/api/licitaciones/export-xlsx")
async def export_licitaciones_xlsx(f: FiltrosQuery, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    q = _apply_filtros(select(Licitacion), f)
    rows = (await db.execute(q.order_by(Licitacion.id))).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Licitaciones"
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(CSV_HEADERS)
    for col_idx, _ in enumerate(CSV_HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(_serialize(r))
    for col_idx, header in enumerate(CSV_HEADERS, start=1):
        max_len = max(len(str(header)), *(len(str(c.value or "")) for c in ws[get_column_letter(col_idx)][1:min(len(ws[get_column_letter(col_idx)]), 100)]))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"licitaciones_{f.anio or 'todos'}_{f.mes or 'todos'}.xlsx"
    from fastapi.responses import Response
    return Response(content=bio.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.get("/api/licitaciones/opciones")
async def opciones_filtro(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    estados = (await db.execute(select(func.distinct(Licitacion.estado)).where(Licitacion.estado != ""))).scalars().all()
    categorias = (await db.execute(select(func.distinct(Licitacion.categoria)).where(Licitacion.categoria != ""))).scalars().all()
    departamentos = (await db.execute(select(func.distinct(Licitacion.departamento)).where(Licitacion.departamento != "").order_by(Licitacion.departamento))).scalars().all()
    metodos = (await db.execute(select(func.distinct(Licitacion.metodo)).where(Licitacion.metodo != ""))).scalars().all()
    modalidades = (await db.execute(select(func.distinct(Licitacion.modalidad)).where(Licitacion.modalidad != ""))).scalars().all()
    return {"estados": list(estados), "categorias": list(categorias),
            "departamentos": list(departamentos), "metodos": list(metodos),
            "modalidades": list(modalidades)}

@app.get("/api/licitaciones/meses")
async def meses_disponibles(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(select(Licitacion.anio, Licitacion.mes).distinct())).all()
    meses = sorted({(r[0], r[1]) for r in rows if r[0]}, key=lambda x: (x[0], x[1]), reverse=True)
    return {"meses": [{"anio": y, "mes": m} for y, m in meses]}

@app.get("/api/licitaciones/opciones")
async def opciones_filtro(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    estados = (await db.execute(select(func.distinct(Licitacion.estado)).where(Licitacion.estado != ""))).scalars().all()
    categorias = (await db.execute(select(func.distinct(Licitacion.categoria)).where(Licitacion.categoria != ""))).scalars().all()
    return {"estados": list(estados), "categorias": list(categorias)}

# ============================================================
# RECURRENTE PAYMENTS
# ============================================================
class CreateCheckoutRequest(BaseModel):
    price_id: str

@app.post("/api/payments/create-checkout")
async def create_checkout(req: CreateCheckoutRequest, request: Request, user: User = Depends(get_current_user)):
    plan_key = None
    for k, v in RECURRENTE_PLANS.items():
        if v["price_id"] == req.price_id: plan_key = k; break
    if not plan_key:
        raise HTTPException(status_code=400, detail="Plan no valido")
    base = settings.FRONTEND_URL.rstrip("/") if settings.FRONTEND_URL else str(request.base_url).rstrip("/")
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(f"{RECURRENTE_API}/checkouts", headers={
                "X-SECRET-KEY": settings.RECURRENTE_SECRET_KEY,
                "Content-Type": "application/json",
            }, json={
                "items": [{"price_id": req.price_id, "quantity": 1}],
                "success_url": f"{base}/suscripcion?success=true",
                "cancel_url": f"{base}/suscripcion?canceled=true",
                "metadata": {"user_id": str(user.id), "plan": plan_key},
                "user_id": user.email,
            })
            data = resp.json()
            if resp.status_code != 201:
                raise HTTPException(status_code=400, detail=data.get("message", "Error al crear checkout"))
            return {"url": data["checkout_url"]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/payments/webhook")
async def recurrent_webhook(request: Request):
    import json, hmac, hashlib, time
    body = await request.body()
    try:
        event = json.loads(body)
    except:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    if settings.RECURRENTE_WEBHOOK_SECRET:
        svix_id = request.headers.get("svix-id", "")
        svix_ts = request.headers.get("svix-timestamp", "")
        svix_sig = request.headers.get("svix-signature", "")
        secret = settings.RECURRENTE_WEBHOOK_SECRET.split("_", 1)[-1]
        import base64
        try:
            secret_bytes = base64.b64decode(secret)
        except Exception:
            secret_bytes = secret.encode()
        signed = f"{svix_id}.{svix_ts}.{body.decode('utf-8')}".encode()
        expected = hmac.new(secret_bytes, signed, hashlib.sha256).digest()
        parts = svix_sig.split(",")
        provided = base64.b64decode(parts[1]) if len(parts) > 1 else b""
        if not hmac.compare_digest(expected, provided):
            raise HTTPException(status_code=400, detail="Firma invalida")

    async def handle_event(data: dict):
        event_type = data.get("event_type", "") or data.get("type", "")
        checkout = data.get("checkout") or {}
        meta = data.get("metadata") or checkout.get("metadata") or {}
        user_id = meta.get("user_id") or checkout.get("user_id")
        plan = meta.get("plan", "pro")
        email = data.get("customer", {}).get("email", "")
        async with async_session() as db:
            user = None
            if user_id:
                try:
                    user = await db.get(User, int(user_id))
                except (TypeError, ValueError):
                    user = None
            if not user and email:
                result = await db.execute(select(User).where(User.email == email))
                user = result.scalar_one_or_none()
            if user:
                if event_type in ("subscription.create", "subscription.reactivate", "subscription.unpause"):
                    user.subscription_plan = plan
                    user.subscription_status = "active"
                    user.keywords_limit = RECURRENTE_PLANS.get(plan, {}).get("keywords", 50)
                elif event_type in ("subscription.cancel", "subscription.pause", "subscription.past_due"):
                    user.subscription_status = "inactive"
                await db.commit()
    try:
        await handle_event(event)
    except Exception as e:
        print(f"Webhook error: {e}")
    return {"ok": True}

@app.get("/api/payments/plans")
async def get_plans():
    return {"plans": [
        {"id": "free", "name": "Free", "price": 0, "keywords": 5, "pipeline": 0, "schedules": 0, "users": 1},
        {"id": "basico",  "name": "B\u00e1sico",  "price": 349, "keywords": 10,  "pipeline": 3,  "schedules": 1, "users": 1,
         "stripe_price_id": RECURRENTE_PLANS["basico"]["price_id"]},
        {"id": "pro",     "name": "Pro",        "price": 599, "keywords": 50,  "pipeline": 30, "schedules": 5, "users": 3,
         "stripe_price_id": RECURRENTE_PLANS["pro"]["price_id"]},
        {"id": "enterprise", "name": "Enterprise","price": 999, "keywords": 999, "pipeline": 999, "schedules": 999, "users": 10,
         "stripe_price_id": RECURRENTE_PLANS["enterprise"]["price_id"]},
    ]}

# ============================================================
# STATIC FILES (Frontend SPA en produccion) - catch-all al FINAL
# ============================================================
STATIC_DIR = Path(__file__).parent.parent / "static"
if STATIC_DIR.exists():
    app.mount("/assets", StaticFiles(directory=str(STATIC_DIR / "assets")), name="assets")
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR / "static"), check_dir=False), name="static")

# ============================================================
# EXTRACTION
# ============================================================
@app.post("/api/extraction/start")
async def start_extraction(user: User = Depends(get_current_user)):
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Solo admin")
    from app.services.extractor_service import run_extraction
    import asyncio
    asyncio.create_task(run_extraction())
    return {"status": "started"}

@app.get("/api/extraction/logs")
async def extraction_logs(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(select(ExtractionLog).order_by(ExtractionLog.created_at.desc()).limit(20))).scalars().all()
    return {"logs": [{"anio": r.anio, "mes": r.mes, "records": r.records_count,
                      "status": r.status, "fecha": str(r.created_at)[:19]} for r in rows]}

@app.get("/api/extraction/status")
async def extraction_status(user: User = Depends(get_current_user)):
    from app.services.extractor_service import next_refresh_at, last_refresh_at
    return {
        "next_refresh_at": next_refresh_at.isoformat() + "Z" if next_refresh_at else None,
        "last_refresh_at": last_refresh_at.isoformat() + "Z" if last_refresh_at else None,
        "interval_hours": 6,
    }

# ============================================================
# ALERTAS (keywords) Y ENVIO POR CORREO
# ============================================================
class AlertRequest(BaseModel):
    keyword: str

@app.get("/api/alerts")
async def mis_alertas(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    oid = _owner_id(user)
    rows = (await db.execute(select(KeywordAlert).where(KeywordAlert.user_id == oid).order_by(KeywordAlert.id))).scalars().all()
    limite = RECURRENTE_PLANS.get(user.subscription_plan, {}).get("keywords", 5)
    return {"alerts": [{"id": a.id, "keyword": a.keyword, "hora_envio": a.hora_envio, "dias_envio": a.dias_envio, "frecuencia": a.frecuencia or "diario", "ultimo_envio": str(a.ultimo_envio)[:19] if a.ultimo_envio else None} for a in rows], "limite": limite}

@app.post("/api/alerts")
async def crear_alerta(req: AlertRequest, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    keyword = req.keyword.strip()
    oid = _owner_id(user)
    if not keyword:
        raise HTTPException(status_code=400, detail="Keyword vacia")
    limite = RECURRENTE_PLANS.get(user.subscription_plan, {}).get("keywords", 5)
    count = await db.scalar(select(func.count()).select_from(KeywordAlert).where(KeywordAlert.user_id == oid))
    if count >= limite:
        raise HTTPException(status_code=400, detail=f"Limite de {limite} keywords alcanzado para tu plan")
    exist = await db.scalar(select(KeywordAlert).where(KeywordAlert.user_id == oid, KeywordAlert.keyword.ilike(keyword)))
    if exist:
        raise HTTPException(status_code=400, detail="Keyword ya registrada")
    alert = KeywordAlert(user_id=oid, keyword=keyword)
    db.add(alert); await db.commit()
    return {"id": alert.id, "keyword": keyword}

@app.delete("/api/alerts/{alert_id}")
async def eliminar_alerta(alert_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    oid = _owner_id(user)
    row = await db.get(KeywordAlert, alert_id)
    if not row or row.user_id != oid:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")
    await db.delete(row); await db.commit()
    return {"ok": True}

@app.post("/api/alerts/{alert_id}/test")
async def test_alerta(alert_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    oid = _owner_id(user)
    alert = await db.get(KeywordAlert, alert_id)
    if not alert or alert.user_id != oid:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")
    ahora = datetime.utcnow()
    rows = (await db.execute(
        select(Licitacion).where(
            Licitacion.titulo.ilike(f"%{alert.keyword}%"),
            Licitacion.anio == ahora.year, Licitacion.mes == ahora.month,
        ).limit(10)
    )).scalars().all()
    if not rows:
        raise HTTPException(status_code=400, detail="No se encontraron coincidencias para esta palabra clave este mes")
    lista = "".join(
        f'<li><a href="https://guatecompras.gt/procesos/{r.nog}" style="color:#1a5fb4">{r.titulo}</a> - Q{float(r.monto or 0):,.0f}</li>'
        for r in rows[:10])
    html = f"""<div style="font-family:Arial;max-width:600px;margin:auto;color:#222">
        <h2 style="color:#1a3a5c">LiciTrackGT - Prueba de alerta</h2>
        <p>Palabra clave: <b>{alert.keyword}</b></p>
        <p>Coincidencias encontradas este mes: <b>{len(rows)}</b></p>
        <ul style="font-size:13px;line-height:1.6">{lista}</ul>
        <p><a href="{settings.FRONTEND_URL}" style="background:#1a3a5c;color:#fff;padding:8px 16px;border-radius:6px;text-decoration:none">Abrir LiciTrackGT</a></p>
    </div>"""
    from app.services.email_service import enviar_correo
    from app.services.whatsapp_service import enviar_whatsapp
    result = {"email": False, "whatsapp": False, "registros": len(rows)}
    try:
        await enviar_correo([user.email], f"LiciTrackGT - Prueba: {alert.keyword}", html)
        result["email"] = True
    except Exception as e:
        result["error"] = str(e)[:100]
    if user.whatsapp_phone:
        try:
            wa_text = f"LiciTrackGT - Prueba\n\n{alert.keyword}: {len(rows)} coincidencias\n" + "\n".join(f"  {r.titulo[:60]}" for r in rows[:5])
            ok = await enviar_whatsapp(user.whatsapp_phone, wa_text)
            result["whatsapp"] = ok
        except Exception as e:
            result["error"] = str(e)[:100]
    return result

class UpdateAlertRequest(BaseModel):
    hora_envio: Optional[int] = None
    dias_envio: Optional[str] = None
    frecuencia: Optional[str] = None

@app.patch("/api/alerts/{alert_id}")
async def actualizar_alerta(alert_id: int, req: UpdateAlertRequest, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    oid = _owner_id(user)
    row = await db.get(KeywordAlert, alert_id)
    if not row or row.user_id != oid:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")
    if req.hora_envio is not None:
        if req.hora_envio < 0 or req.hora_envio > 23:
            raise HTTPException(status_code=400, detail="Hora debe ser 0-23")
        row.hora_envio = req.hora_envio
    if req.dias_envio is not None:
        row.dias_envio = req.dias_envio
    if req.frecuencia is not None:
        row.frecuencia = req.frecuencia
    await db.commit()
    return {"ok": True, "id": row.id, "keyword": row.keyword, "hora_envio": row.hora_envio, "dias_envio": row.dias_envio, "frecuencia": row.frecuencia or "diario"}

class PipelineRequest(BaseModel):
    nog: str
    titulo: str = ""
    entidad: str = ""
    monto: float = 0
    fecha_publicacion: Optional[str] = None

@app.get("/api/pipeline")
async def mi_pipeline(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(select(PipelineItem).where(PipelineItem.user_id == _owner_id(user)).order_by(PipelineItem.updated_at.desc()))).scalars().all()
    limite = RECURRENTE_PLANS.get(user.subscription_plan, {}).get("pipeline", 0)
    return {"items": [{"id": p.id, "nog": p.nog, "titulo": p.titulo, "entidad": p.entidad, "monto": p.monto,
                        "fecha_publicacion": str(p.fecha_publicacion) if p.fecha_publicacion else "",
                        "etapa": p.etapa, "fecha_presentacion": str(p.fecha_presentacion) if p.fecha_presentacion else "",
                        "monto_propuesto": p.monto_propuesto or 0, "probabilidad": p.probabilidad or 0,
                        "notas": p.notas or ""} for p in rows], "limite": limite}

@app.post("/api/pipeline")
async def agregar_pipeline(req: PipelineRequest, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    limite = RECURRENTE_PLANS.get(user.subscription_plan, {}).get("pipeline", 0)
    if limite <= 0:
        raise HTTPException(status_code=403, detail="Pipeline disponible en plan Basico o superior")
    count = await db.scalar(select(func.count()).select_from(PipelineItem).where(PipelineItem.user_id == _owner_id(user)))
    if count >= limite:
        raise HTTPException(status_code=400, detail=f"Limite de {limite} items alcanzado para tu plan")
    exist = await db.scalar(select(PipelineItem).where(PipelineItem.user_id == _owner_id(user), PipelineItem.nog == req.nog))
    if exist:
        raise HTTPException(status_code=400, detail="Este NOG ya esta en tu pipeline")
    fp = None
    if req.fecha_publicacion:
        try: fp = date.fromisoformat(req.fecha_publicacion)
        except: fp = None
    p = PipelineItem(user_id=_owner_id(user), nog=req.nog, titulo=req.titulo, entidad=req.entidad,
                     monto=req.monto, fecha_publicacion=fp, etapa="deteccion")
    db.add(p); await db.commit(); await db.refresh(p)
    return {"id": p.id, "nog": p.nog, "etapa": p.etapa}

class UpdatePipelineRequest(BaseModel):
    etapa: Optional[str] = None
    fecha_presentacion: Optional[str] = None
    monto_propuesto: Optional[float] = None
    probabilidad: Optional[int] = None
    notas: Optional[str] = None

@app.patch("/api/pipeline/{item_id}")
async def actualizar_pipeline(item_id: int, req: UpdatePipelineRequest, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    p = await db.get(PipelineItem, item_id)
    if not p or p.user_id != _owner_id(user):
        raise HTTPException(status_code=404, detail="Item no encontrado")
    if req.etapa and req.etapa not in PIPELINE_ETAPAS:
        raise HTTPException(status_code=400, detail=f"Etapa invalida. Usa: {', '.join(PIPELINE_ETAPAS)}")
    if req.etapa: p.etapa = req.etapa
    if req.fecha_presentacion is not None:
        try: p.fecha_presentacion = date.fromisoformat(req.fecha_presentacion) if req.fecha_presentacion else None
        except: pass
    if req.monto_propuesto is not None: p.monto_propuesto = req.monto_propuesto
    if req.probabilidad is not None: p.probabilidad = max(0, min(100, req.probabilidad))
    if req.notas is not None: p.notas = req.notas
    from datetime import datetime
    p.updated_at = datetime.utcnow()
    await db.commit()
    return {"ok": True, "id": p.id, "etapa": p.etapa}

@app.delete("/api/pipeline/{item_id}")
async def eliminar_pipeline(item_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    p = await db.get(PipelineItem, item_id)
    if not p or p.user_id != _owner_id(user):
        raise HTTPException(status_code=404, detail="Item no encontrado")
    await db.delete(p); await db.commit()
    return {"ok": True}

@app.get("/api/pipeline/deadlines")
async def pipeline_deadlines(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(
        select(PipelineItem).where(PipelineItem.user_id == _owner_id(user), PipelineItem.fecha_presentacion != None)
        .order_by(PipelineItem.fecha_presentacion).limit(10)
    )).scalars().all()
    hoy = date.today()
    return {"deadlines": [{"nog": p.nog, "titulo": p.titulo, "etapa": p.etapa,
                            "fecha_presentacion": str(p.fecha_presentacion),
                            "dias_faltan": (p.fecha_presentacion - hoy).days} for p in rows if p.fecha_presentacion and p.fecha_presentacion >= hoy]}

class ScheduleRequest(BaseModel):
    hora: int = 8
    dias: str = "1,2,3,4,5"
    recipients: str = ""
    keywords: str = ""
    anio: Optional[int] = None
    mes: Optional[int] = None

@app.get("/api/scheduled-reports")
async def mis_reportes(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(select(ScheduledReport).where(ScheduledReport.user_id == _owner_id(user)).order_by(ScheduledReport.id))).scalars().all()
    limite = RECURRENTE_PLANS.get(user.subscription_plan, {}).get("schedules", 0)
    return {"items": [{"id": r.id, "hora": r.hora, "dias": r.dias, "enabled": r.enabled,
                        "recipients": r.recipients, "keywords": r.keywords,
                        "anio": r.anio, "mes": r.mes,
                        "ultimo_envio": str(r.ultimo_envio)[:19] if r.ultimo_envio else None} for r in rows], "limite": limite}

@app.post("/api/scheduled-reports")
async def crear_reporte(req: ScheduleRequest, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    limite = RECURRENTE_PLANS.get(user.subscription_plan, {}).get("schedules", 0)
    if limite <= 0:
        raise HTTPException(status_code=403, detail="Reportes programados disponibles en plan Basico o superior")
    count = await db.scalar(select(func.count()).select_from(ScheduledReport).where(ScheduledReport.user_id == _owner_id(user)))
    if count >= limite:
        raise HTTPException(status_code=400, detail=f"Limite de {limite} reportes alcanzado para tu plan")
    r = ScheduledReport(user_id=_owner_id(user), hora=req.hora, dias=req.dias, recipients=req.recipients.strip(),
                        keywords=req.keywords.strip(), anio=req.anio, mes=req.mes)
    db.add(r); await db.commit(); await db.refresh(r)
    return {"id": r.id, "hora": r.hora}

@app.patch("/api/scheduled-reports/{report_id}")
async def toggle_reporte(report_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    r = await db.get(ScheduledReport, report_id)
    if not r or r.user_id != _owner_id(user):
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    r.enabled = not r.enabled
    await db.commit()
    return {"id": r.id, "enabled": r.enabled}

@app.delete("/api/scheduled-reports/{report_id}")
async def eliminar_reporte(report_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    r = await db.get(ScheduledReport, report_id)
    if not r or r.user_id != _owner_id(user):
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    await db.delete(r); await db.commit()
    return {"ok": True}

class InviteRequest(BaseModel):
    email: str
    name: str = ""

@app.get("/api/team/members")
async def miembros_equipo(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    owner_id = user.main_user_id or user.id
    owner = await db.get(User, owner_id)
    max_users = RECURRENTE_PLANS.get(owner.subscription_plan, {}).get("users", 1)
    members = (await db.execute(select(User).where(
        (User.id == owner_id) | (User.main_user_id == owner_id)
    ))).scalars().all()
    return {"members": [{"id": m.id, "email": m.email, "name": m.name, "plan": m.subscription_plan} for m in members], "max_users": max_users, "owner_id": owner_id}

@app.post("/api/team/invite")
async def invitar_miembro(req: InviteRequest, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    owner_id = user.main_user_id or user.id
    owner = await db.get(User, owner_id)
    max_users = RECURRENTE_PLANS.get(owner.subscription_plan, {}).get("users", 1)
    current = await db.scalar(select(func.count()).select_from(User).where(
        (User.id == owner_id) | (User.main_user_id == owner_id)
    ))
    if current >= max_users:
        raise HTTPException(status_code=400, detail=f"Limite de {max_users} usuarios alcanzado")
    existing = await db.scalar(select(User).where(User.email == req.email))
    if existing:
        if existing.main_user_id:
            raise HTTPException(status_code=400, detail="Usuario ya pertenece a otro equipo")
        existing.main_user_id = owner_id
        existing.subscription_plan = owner.subscription_plan
        existing.subscription_status = "active"
        existing.keywords_limit = owner.keywords_limit
        await db.commit()
        _send_invite_email(req.email, req.name, owner.email, None)
        return {"id": existing.id, "email": existing.email}
    import secrets
    temp_pass = secrets.token_urlsafe(8)
    new_user = User(email=req.email, name=req.name, password_hash=hash_password(temp_pass),
                    main_user_id=owner_id, subscription_plan=owner.subscription_plan,
                    subscription_status="active", keywords_limit=owner.keywords_limit)
    db.add(new_user); await db.commit(); await db.refresh(new_user)
    _send_invite_email(req.email, req.name, owner.email, temp_pass)
    return {"id": new_user.id, "email": new_user.email, "temp_password": temp_pass}

@app.delete("/api/team/members/{member_id}")
async def eliminar_miembro(member_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    owner_id = user.main_user_id or user.id
    member = await db.get(User, member_id)
    if not member or member.main_user_id != owner_id:
        raise HTTPException(status_code=404, detail="Miembro no encontrado")
    member.main_user_id = None
    await db.commit()
    return {"ok": True}

class EnviarCorreoRequest(BaseModel):
    destinatario: str = ""
    asunto: Optional[str] = None

@app.post("/api/licitaciones/enviar")
async def enviar_resultados(f: FiltrosQuery, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if user.subscription_plan == "free":
        raise HTTPException(status_code=403, detail="Funcion disponible en plan Basico o superior")
    q = _apply_filtros(select(Licitacion), f).limit(1000)
    rows = (await db.execute(q.order_by(Licitacion.fecha_publicacion.desc()))).scalars().all()
    if not rows:
        raise HTTPException(status_code=400, detail="No hay resultados para enviar")
    raw = (f.destinatarios or f.destinatario or user.email).strip()
    destinos = [d.strip() for d in raw.replace(";", ",").split(",") if d.strip()]
    validos = [d for d in destinos if "@" in d and "." in d]
    if not validos:
        raise HTTPException(status_code=400, detail="Correo(s) invalido(s)")
    titulo_filtro = f.texto or f"filtros ({len(rows)} resultados)"

    import io, csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(CSV_HEADERS)
    for r in rows:
        writer.writerow(_serialize(r))
    csv_bytes = buf.getvalue().encode("utf-8-sig")

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Licitaciones"
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(CSV_HEADERS)
    for col_idx, _ in enumerate(CSV_HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(_serialize(r))
    for col_idx, header in enumerate(CSV_HEADERS, start=1):
        col = get_column_letter(col_idx)
        max_len = max(len(str(header)), *(len(str(c.value or "")) for c in ws[col][1:min(len(ws[col]), 100)]))
        ws.column_dimensions[col].width = min(max_len + 2, 60)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    xbio = BytesIO()
    wb.save(xbio); xbio.seek(0)
    xlsx_bytes = xbio.getvalue()

    fila = "".join(
        f'<tr><td>{r.nog}</td><td>{r.fecha_publicacion}</td><td><a href="https://guatecompras.gt/procesos/{r.nog}" style="color:#1a5fb4">{r.titulo}</a></td>'
        f'<td style="text-align:right">Q{float(r.monto or 0):,.0f}</td><td>{r.entidad_compradora}</td></tr>'
        for r in rows[:80])
    html = f"""<div style="font-family:Arial;max-width:700px;margin:auto;color:#222">
        <h2 style="color:#1a3a5c">LiciTrackGT - {len(rows)} licitaciones</h2>
        <p>Filtro: <b>{titulo_filtro}</b></p>
        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;width:100%;font-size:13px">
            <tr style="background:#1a3a5c;color:#fff;text-align:left">
                <th>NOG</th><th>Fecha</th><th>Titulo</th><th>Monto</th><th>Entidad</th></tr>
            {fila}
        </table>
        <p style="margin-top:16px">Adjunto: <b>XLSX</b> con el detalle completo de {len(rows)} licitaciones.</p>
        <p><a href="{settings.FRONTEND_URL}" style="background:#1a3a5c;color:#fff;padding:8px 16px;border-radius:6px;text-decoration:none">Abrir LiciTrackGT</a></p>
        <p style="color:#888;font-size:12px">Recibes este correo por solicitud en tu cuenta.</p></div>"""

    from app.services.email_service import enviar_correo
    try:
        await enviar_correo(validos, f"LiciTrackGT: {len(rows)} licitaciones - {titulo_filtro[:50]}",
                            html, (f"licitaciones_{f.anio or 'todos'}_{f.mes or 'todos'}.xlsx",
                                   xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar correo: {str(e)[:200]}")
    return {"ok": True, "enviado_a": validos, "registros": len(rows)}

# ============================================================
# ADMIN
# ============================================================
class AdminUpdateUserRequest(BaseModel):
    plan: str
    keywords_limit: int

@app.get("/api/admin/usuarios")
async def admin_list_users(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Solo admin")
    result = await db.execute(select(User).order_by(User.id))
    users = result.scalars().all()
    return [{
        "id": u.id, "email": u.email, "name": u.name,
        "plan": u.subscription_plan or "free",
        "status": u.subscription_status or "inactive",
        "whatsapp_phone": u.whatsapp_phone or "",
        "created_at": str(u.created_at)[:19] if u.created_at else None,
    } for u in users]

@app.get("/api/admin/stats")
async def admin_stats(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Solo admin")
    total_users = await db.scalar(select(func.count(User.id)))
    paying_users = await db.scalar(
        select(func.count(User.id)).where(User.subscription_status == "active", User.subscription_plan != "free")
    )
    total_alerts = await db.scalar(select(func.count(KeywordAlert.id)))
    total_pipeline = await db.scalar(select(func.count(PipelineItem.id)))
    total_scheduled = await db.scalar(select(func.count(ScheduledReport.id)))
    total_licitaciones = await db.scalar(select(func.count(Licitacion.id)))
    paying_users_list = (await db.execute(
        select(User).where(User.subscription_status == "active", User.subscription_plan != "free")
    )).scalars().all()
    mrr = sum(RECURRENTE_PLANS.get(u.subscription_plan, {}).get("price", 0) for u in paying_users_list)
    return {
        "total_users": total_users or 0,
        "paying_users": paying_users or 0,
        "total_alerts_count": total_alerts or 0,
        "total_pipeline_count": total_pipeline or 0,
        "total_scheduled_reports_count": total_scheduled or 0,
        "total_licitaciones": total_licitaciones or 0,
        "mrr": mrr,
    }

@app.patch("/api/admin/usuarios/{user_id}")
async def admin_update_user(user_id: int, req: AdminUpdateUserRequest, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Solo admin")
    target = await db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    target.subscription_plan = req.plan
    target.keywords_limit = req.keywords_limit
    await db.commit()
    return {"ok": True, "id": target.id, "plan": target.subscription_plan, "keywords_limit": target.keywords_limit}

# ============================================================
# STARTUP
# ============================================================
@app.get("/api/debug/network", include_in_schema=False)
async def debug_network(user: User = Depends(get_current_user)):
    import socket
    result = {"smtp_host": settings.SMTP_HOST, "smtp_port": settings.SMTP_PORT,
              "smtp_user_configured": bool(settings.SMTP_USER)}
    for host, port in [("smtp.gmail.com", 587), ("smtp.gmail.com", 465), ("ocds.guatecompras.gt", 443)]:
        try:
            infos = socket.getaddrinfo(host, port, socket.AF_UNSPEC, socket.SOCK_STREAM)
            result[f"{host}:{port}_dns"] = [i[4][0] for i in infos]
            s = socket.create_connection((host, port), timeout=8)
            s.close()
            result[f"{host}:{port}_connect"] = "OK"
        except Exception as e:
            result[f"{host}:{port}_connect"] = f"{type(e).__name__}: {e}"
    try:
        s = socket.create_connection(("142.250.141.108", 587), timeout=8)
        s.close()
        result["smtp_ipv4_connect"] = "OK"
    except Exception as e:
        result["smtp_ipv4_connect"] = f"{type(e).__name__}: {e}"
    return result

@app.on_event("startup")
async def startup():
    await init_db()
    async with async_session() as db:
        try:
            await db.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS whatsapp_phone VARCHAR(30)"))
            await db.commit()
        except Exception: pass
        try:
            await db.execute(text("ALTER TABLE keyword_alerts ADD COLUMN IF NOT EXISTS hora_envio INTEGER"))
            await db.commit()
        except Exception: pass
        try:
            await db.execute(text("ALTER TABLE keyword_alerts ADD COLUMN IF NOT EXISTS dias_envio VARCHAR(50) DEFAULT '1,2,3,4,5'"))
            await db.commit()
        except Exception: pass
        try:
            await db.execute(text("ALTER TABLE keyword_alerts ADD COLUMN IF NOT EXISTS ultimo_envio TIMESTAMPTZ"))
            await db.commit()
        except Exception: pass
        try:
            await db.execute(text("ALTER TABLE keyword_alerts ADD COLUMN IF NOT EXISTS frecuencia VARCHAR(20) DEFAULT 'diario'"))
            await db.commit()
        except Exception: pass
        try:
            await db.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS main_user_id INTEGER REFERENCES users(id)"))
            await db.commit()
        except Exception: pass
        result = await db.execute(select(User).where(User.email == "totalappgt@gmail.com"))
        if not result.scalar_one_or_none():
            admin = User(email="totalappgt@gmail.com", password_hash=hash_password("admintotal"),
                         name="Admin", is_admin=True, is_active=True, subscription_plan="enterprise",
                         subscription_status="active", keywords_limit=999)
            db.add(admin); await db.commit()
            print("Admin creado: totalappgt@gmail.com / admintotal")
    import asyncio
    from app.services.extractor_service import background_auto_refresh
    asyncio.create_task(background_auto_refresh())
    print("LiciTrackGT API iniciada")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=True)

from sqlalchemy.ext.asyncio import async_sessionmaker
from app.database import async_session

# ============================================================
# SPA CATCH-ALL (debe ir despues de todas las rutas API)
# ============================================================
if STATIC_DIR.exists():
    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_spa(full_path: str):
        if full_path.startswith("api/") or full_path.startswith("docs") or full_path.startswith("openapi"):
            from fastapi.responses import JSONResponse
            return JSONResponse({"detail": "Not Found"}, status_code=404)
        from fastapi.responses import FileResponse
        index = STATIC_DIR / "index.html"
        if not index.exists():
            return JSONResponse({"detail": "Frontend no construido"}, status_code=500)
        return FileResponse(str(index))
