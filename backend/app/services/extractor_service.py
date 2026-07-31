import httpx
from datetime import date, datetime, timedelta
from app.database import async_session
from app.models import Licitacion, ExtractionLog, KeywordAlert, ScheduledReport, PipelineItem, User
from app.config import settings

OCDS_BASE = "https://ocds.guatecompras.gt"

next_refresh_at = None
last_refresh_at = None

async def obtener_meses_disponibles():
    async with httpx.AsyncClient(timeout=30) as cl:
        r = await cl.get(f"{OCDS_BASE}/v1/releases/bulk")
        if r.status_code != 200: return []
        data = r.json()
        return data.get("months", [])

def parse_releases(data: dict) -> list:
    releases = []
    for seccion in ["tenders", "awards", "contracts"]:
        for item in data.get(seccion, []):
            r = item.get("release", item)
            releases.append(r)
    return releases

async def run_extraction(anio: int = None, mes: int = None):
    meses = await obtener_meses_disponibles()
    if not meses: return
    if anio: meses = [m for m in meses if m.get("year") == anio or m.get("anio") == anio]
    if mes: meses = [m for m in meses if m.get("month") == mes or m.get("mes") == mes]
    async with httpx.AsyncClient(timeout=120) as cl:
        for m in meses:
            y = m.get("year", m.get("anio"))
            mo = m.get("month", m.get("mes"))
            key = f"{y}-{mo}"
            try:
                resp = await cl.get(f"{OCDS_BASE}/v1/releases/bulk", params={"year": y, "month": mo})
                if resp.status_code != 200: continue
                data = resp.json()
                items = data if isinstance(data, list) else data.get("data", data.get("releases", []))
                if not items: continue
                count = 0
                async with async_session() as db:
                    for item in items:
                        release = item if isinstance(item, dict) else {}
                        tender = release.get("tender") or {}
                        buyer = release.get("buyer") or {}
                        nog = tender.get("id", "") or release.get("ocid", "")
                        if not nog: continue
                        try:
                            fp = (release.get("date", "") or "")[:10]
                            fecha_pub = date.fromisoformat(fp) if fp else None
                        except: fecha_pub = None
                        from sqlalchemy.dialects.postgresql import insert as pg_insert
                        from app.models import Licitacion as _L
                        stmt = pg_insert(_L).values(
                            nog=nog, ocid=release.get("ocid", ""),
                            fecha_publicacion=fecha_pub,
                            titulo=tender.get("title", ""),
                            entidad_compradora=buyer.get("name", "") or tender.get("procuringEntity", {}).get("name", ""),
                            monto=float(tender.get("value", {}).get("amount", 0) or 0),
                            moneda=tender.get("value", {}).get("currency", "GTQ"),
                            estado=tender.get("status", ""),
                            categoria=tender.get("mainProcurementCategory", ""),
                            metodo=tender.get("procurementMethod", ""),
                            modalidad=tender.get("procurementMethodDetails", ""),
                            anio=y, mes=mo,
                        ).on_conflict_do_update(
                            index_elements=[_L.nog],
                            set_={
                                "fecha_publicacion": stmt.excluded.fecha_publicacion,
                                "titulo": stmt.excluded.titulo,
                                "entidad_compradora": stmt.excluded.entidad_compradora,
                                "monto": stmt.excluded.monto,
                                "estado": stmt.excluded.estado,
                                "categoria": stmt.excluded.categoria,
                                "metodo": stmt.excluded.metodo,
                                "modalidad": stmt.excluded.modalidad,
                                "anio": stmt.excluded.anio,
                                "mes": stmt.excluded.mes,
                            }
                        )
                        await db.execute(stmt); count += 1
                        if count % 200 == 0: await db.flush()
                    await db.commit()
                    log = ExtractionLog(anio=y, mes=mo, records_count=count, status="completed")
                    db.add(log); await db.commit()
            except Exception as e:
                async with async_session() as db:
                    log = ExtractionLog(anio=y, mes=mo, records_count=0, status=f"error: {str(e)[:100]}")
                    db.add(log); await db.commit()

async def refresh_ultimo_mes():
    import asyncio
    from datetime import datetime
    ahora = datetime.utcnow()
    try:
        await run_extraction(anio=ahora.year, mes=ahora.month)
    except Exception as e:
        print(f"Auto-refresh fallo: {e}")

async def procesar_alertas():
    from sqlalchemy import select
    from app.models import User
    hora_actual = datetime.utcnow().hour
    dia_actual = str(datetime.utcnow().isoweekday())
    ahora = datetime.utcnow()
    async with async_session() as db:
        alertas = (await db.execute(select(KeywordAlert))).scalars().all()
        user_kw_map = {}
        alerts_to_update = []
        for a in alertas:
            if a.hora_envio is not None and a.hora_envio != hora_actual:
                continue
            if a.dias_envio and dia_actual not in a.dias_envio.split(","):
                continue
            if a.hora_envio is not None:
                if a.ultimo_envio and a.ultimo_envio.date() == ahora.date():
                    continue
            else:
                if a.ultimo_envio and a.ultimo_envio.date() == ahora.date():
                    continue
            user_kw_map.setdefault(a.user_id, []).append(a.keyword)
            alerts_to_update.append(a)
        if not user_kw_map:
            return
        hora_str = f" {hora_actual}:00" if any(a.hora_envio is not None for a in alertas) else ""
        ahora = datetime.utcnow()
        user_rows = (await db.execute(select(User).where(User.id.in_(list(user_kw_map.keys()))))).scalars().all()
        for u in user_rows:
            keywords = user_kw_map[u.id]
            if u.subscription_plan in ("free",) or u.subscription_status != "active":
                continue
            matches = []
            for kw in keywords:
                q = select(Licitacion).where(
                    Licitacion.titulo.ilike(f"%{kw}%"),
                    Licitacion.anio == ahora.year,
                    Licitacion.mes == ahora.month,
                ).limit(20)
                for r in (await db.execute(q)).scalars().all():
                    matches.append({"keyword": kw, "nog": r.nog, "titulo": r.titulo,
                                    "monto": r.monto or 0, "fecha": str(r.fecha_publicacion or "")})
            if matches:
                lista = "".join(
                    f'<li><b>{m["keyword"]}</b> - <a href="https://guatecompras.gt/procesos/{m["nog"]}">{m["titulo"]}</a> - Q{float(m["monto"]):,.0f} ({m["fecha"]})</li>'
                    for m in matches[:50])
                html = f"""<div style="font-family:Arial;max-width:600px;margin:auto">
                    <h2 style="color:#1a3a5c">LiciTrackGT - {len(matches)} alertas{hora_str}</h2>
                    <p>Nuevas licitaciones que coinciden con tus keywords:</p>
                    <ul>{lista}</ul>
                    <p><a href="{settings.FRONTEND_URL}">Abrir LiciTrackGT</a></p>
                    <p style="color:#888;font-size:12px">Para dejar de recibir alertas, elimina la keyword en tu panel.</p></div>"""
                from app.services.email_service import enviar_correo
                from app.services.whatsapp_service import enviar_whatsapp
                try:
                    await enviar_correo([u.email], f"LiciTrackGT: {len(matches)} nuevas coincidencias", html)
                except Exception as e:
                    print(f"Error alerta para {u.email}: {e}")
                if u.whatsapp_phone:
                    wa_text = f"LiciTrackGT: {len(matches)} coincidencias{hora_str}\n" + "\n".join(
                        f"  {m['keyword']}: {m['titulo'][:60]} Q{float(m['monto']):,.0f}"
                        for m in matches[:10]
                    ) + f"\n\nVer mas: {settings.FRONTEND_URL}"
                    try:
                        await enviar_whatsapp(u.whatsapp_phone, wa_text)
                    except Exception as e:
                        print(f"Error WhatsApp para {u.email}: {e}")
        for a in alerts_to_update:
            a.ultimo_envio = datetime.utcnow()
        await db.commit()

async def procesar_scheduled_reports():
    from sqlalchemy import select
    from app.models import User
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    hora_actual = datetime.utcnow().hour
    dia_actual = str(datetime.utcnow().isoweekday())
    async with async_session() as db:
        reports = (await db.execute(
            select(ScheduledReport).where(ScheduledReport.enabled == True)
        )).scalars().all()
        for rep in reports:
            if rep.hora != hora_actual:
                continue
            if rep.dias and dia_actual not in rep.dias.split(","):
                continue
            if rep.ultimo_envio and rep.ultimo_envio.date() == datetime.utcnow().date():
                continue
            user = await db.get(User, rep.user_id)
            if not user or user.subscription_plan == "free" or user.subscription_status != "active":
                continue
            kws = [k.strip() for k in rep.keywords.split(",") if k.strip()] if rep.keywords else []
            q = select(Licitacion)
            if rep.anio: q = q.where(Licitacion.anio == rep.anio)
            if rep.mes: q = q.where(Licitacion.mes == rep.mes)
            if kws:
                from sqlalchemy import or_
                q = q.where(or_(*[Licitacion.titulo.ilike(f"%{kw}%") for kw in kws]))
            rows = (await db.execute(q.order_by(Licitacion.fecha_publicacion.desc()).limit(500))).scalars().all()
            if not rows:
                continue
            recips = [r.strip() for r in (rep.recipients or user.email).replace(";", ",").split(",") if r.strip() and "@" in r]
            if not recips:
                continue
            headers = ["NOG", "OCID", "Fecha", "Titulo", "Entidad", "Monto", "Moneda", "Estado", "Categoria", "Metodo", "Modalidad", "Departamento"]
            def _ser(r):
                return [r.nog, r.ocid, str(r.fecha_publicacion) if r.fecha_publicacion else "",
                        r.titulo or "", r.entidad_compradora or "", r.monto or 0, r.moneda or "GTQ",
                        r.estado or "", r.categoria or "", r.metodo or "", r.modalidad or "", r.departamento or ""]
            wb = Workbook()
            ws = wb.active
            ws.title = "Licitaciones"
            hfill = PatternFill("solid", fgColor="1A3A5C")
            hfont = Font(color="FFFFFF", bold=True)
            ws.append(headers)
            for ci, _ in enumerate(headers, start=1):
                c = ws.cell(row=1, column=ci)
                c.fill = hfill; c.font = hfont; c.alignment = Alignment(horizontal="center")
            for r in rows:
                ws.append(_ser(r))
            for ci, h in enumerate(headers, start=1):
                col = get_column_letter(ci)
                mx = max(len(str(h)), *(len(str(c.value or "")) for c in ws[col][1:min(len(ws[col]), 100)]))
                ws.column_dimensions[col].width = min(mx + 2, 60)
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"
            xbio = BytesIO(); wb.save(xbio); xbio.seek(0)
            keyword_text = rep.keywords or "sin filtro"
            html = f"""<div style="font-family:Arial;max-width:600px;margin:auto;color:#222">
                <h2 style="color:#1a3a5c">LiciTrackGT - Reporte programado</h2>
                <p><b>{len(rows)} licitaciones</b> para: {keyword_text}</p>
                <p style="margin:16px 0">Adjunto: <b>XLSX</b> con el detalle completo.</p>
                <p><a href="{settings.FRONTEND_URL}" style="background:#1a3a5c;color:#fff;padding:8px 16px;border-radius:6px;text-decoration:none">Abrir LiciTrackGT</a></p>
                <p style="color:#888;font-size:12px">Reporte programado desde tu cuenta. Puedes desactivarlo en Alertas.</p></div>"""
            from app.services.email_service import enviar_correo
            from app.services.whatsapp_service import enviar_whatsapp
            try:
                await enviar_correo(recips, f"LiciTrackGT: reporte programado - {keyword_text[:50]}",
                                    html, (f"reporte_{rep.anio or 'todo'}_{rep.mes or 'todo'}.xlsx",
                                           xbio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
                rep.ultimo_envio = datetime.utcnow()
                await db.commit()
            except Exception as e:
                print(f"Error reporte programado id={rep.id}: {e}")
            if user.whatsapp_phone:
                try:
                    await enviar_whatsapp(user.whatsapp_phone,
                                          f"LiciTrackGT Reporte: {len(rows)} licitaciones para '{keyword_text[:40]}'\nAdjunto XLSX en tu correo.\n{settings.FRONTEND_URL}")
                except Exception as e:
                    print(f"Error WhatsApp reporte id={rep.id}: {e}")

async def procesar_deadline_alerts():
    from sqlalchemy import select
    hoy = date.today()
    async with async_session() as db:
        items = (await db.execute(
            select(PipelineItem).where(
                PipelineItem.fecha_presentacion != None,
                PipelineItem.etapa.notin_(["ganada", "perdida"]),
            ).order_by(PipelineItem.fecha_presentacion)
        )).scalars().all()
        for p in items:
            if not p.fecha_presentacion:
                continue
            dias = (p.fecha_presentacion - hoy).days
            if dias not in (0, 1, 2, 3):
                continue
            user = await db.get(User, p.user_id)
            if not user or user.subscription_plan == "free" or user.subscription_status != "active":
                continue
            label = "HOY" if dias == 0 else f"{dias} dia(s)" if dias <= 1 else f"{dias} dias"
            asunto = f"ALERTA: Presentacion en {label} - {p.titulo[:40]}"
            html = f"""<div style="font-family:Arial;max-width:600px;margin:auto;color:#222">
                <h2 style="color:#c00">LiciTrackGT - Recordatorio de presentacion</h2>
                <p style="font-size:16px"><b>{p.titulo}</b></p>
                <p>NOG: <a href="https://guatecompras.gt/procesos/{p.nog}">{p.nog}</a></p>
                <p>Fecha de presentacion: <b style="color:#c00">{p.fecha_presentacion}</b> ({label})</p>
                <p>Entidad: {p.entidad or 'N/A'} | Etapa: {p.etapa}</p>
                {f'<p>Monto propuesto: Q{float(p.monto_propuesto or 0):,.0f}</p>' if p.monto_propuesto else ''}
                <p><a href="{settings.FRONTEND_URL}" style="background:#1a3a5c;color:#fff;padding:8px 16px;border-radius:6px;text-decoration:none">Ir al Pipeline</a></p>
            </div>"""
            from app.services.email_service import enviar_correo
            from app.services.whatsapp_service import enviar_whatsapp
            try:
                await enviar_correo([user.email], asunto, html)
            except Exception as e:
                print(f"Error deadline alert email {user.email}: {e}")
            if user.whatsapp_phone:
                try:
                    await enviar_whatsapp(user.whatsapp_phone,
                                          f"LiciTrackGT: Presentacion en {label}\n{p.titulo[:80]}\nNOG: {p.nog}\nEntidad: {p.entidad or 'N/A'}\nVer: https://guatecompras.gt/procesos/{p.nog}")
                except Exception as e:
                    print(f"Error deadline alert whatsapp {user.email}: {e}")

async def background_auto_refresh():
    import asyncio
    global next_refresh_at, last_refresh_at
    last_refresh_time = datetime.utcnow() - timedelta(hours=7)
    while True:
        ahora = datetime.utcnow()
        if (ahora - last_refresh_time).total_seconds() >= 6 * 3600:
            try:
                await refresh_ultimo_mes()
                last_refresh_at = datetime.utcnow()
                last_refresh_time = last_refresh_at
            except Exception as e:
                print(f"Auto-refresh error: {e}")
        next_refresh_at = last_refresh_time + timedelta(hours=6)
        try:
            await procesar_alertas()
        except Exception as e:
            print(f"Alertas error: {e}")
        try:
            await procesar_scheduled_reports()
        except Exception as e:
            print(f"Scheduled reports error: {e}")
        try:
            await procesar_deadline_alerts()
        except Exception as e:
            print(f"Deadline alerts error: {e}")
        await asyncio.sleep(900)
